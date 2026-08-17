from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from hwscan.domain.models import InventoryReport
from hwscan.domain.normalization import excel_safe_text

HEADER_FILL = PatternFill("solid", fgColor="17324D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="17324D")


def write_excel(report: InventoryReport, path: Path) -> None:
    data = report.as_dict()
    workbook = Workbook()
    active = workbook.active
    if active is not None:
        workbook.remove(active)

    summary = workbook.create_sheet("Summary")
    summary.append(["HWScan USB hardware inventory"])
    summary["A1"].font = TITLE_FONT
    summary.append(["Report ID", report.report_id])
    summary.append(["Created (UTC)", report.created_at])
    summary.append(["Asset tag", excel_safe_text(report.operator_input.asset_tag)])
    summary.append(["Operator", excel_safe_text(report.operator_input.operator)])
    summary.append(["Location", excel_safe_text(report.operator_input.location)])
    summary.append(["Notes", excel_safe_text(report.operator_input.notes)])
    summary.append([])
    _append_mapping(summary, "System", data["system"])
    _append_mapping(summary, "Boot", data["boot"])
    if report.cpus:
        _append_mapping(summary, "CPU", report.cpus[0])
    _append_mapping(summary, "Memory", {k: v for k, v in report.memory.items() if k != "modules"})
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 65

    _add_rows_sheet(workbook, "Memory", report.memory.get("modules", []), "MemoryTable")
    _add_rows_sheet(workbook, "Storage", report.storage, "StorageTable")
    _add_rows_sheet(workbook, "Graphics", report.graphics, "GraphicsTable")
    _add_rows_sheet(workbook, "Network", report.network, "NetworkTable")
    _add_rows_sheet(workbook, "Battery", report.batteries, "BatteryTable")
    _add_rows_sheet(workbook, "Displays", report.displays, "DisplaysTable")
    device_rows = [{"category": "audio", **item} for item in report.audio] + [
        {"category": "usb", **item} for item in report.usb_devices
    ]
    _add_rows_sheet(workbook, "Devices", device_rows, "DevicesTable")
    _add_rows_sheet(
        workbook,
        "Diagnostics",
        [
            {
                "severity": item.severity.value,
                "domain": item.domain,
                "code": item.code,
                "message": item.message,
                "device_ref": item.device_ref,
                "source_ref": item.source_ref,
            }
            for item in report.diagnostics
        ],
        "DiagnosticsTable",
    )
    metadata_rows = [
        {"key": "schema_version", "value": report.schema_version},
        {"key": "report_id", "value": report.report_id},
        {"key": "created_at", "value": report.created_at},
        *({"key": f"scanner.{key}", "value": value} for key, value in data["scanner"].items()),
        *(
            {
                "key": f"source.{source.id}",
                "value": f"{source.tool} | {source.status.value} | sha256={source.raw_sha256}",
            }
            for source in report.sources
        ),
    ]
    _add_rows_sheet(workbook, "Metadata", metadata_rows, "MetadataTable")

    workbook.properties.title = "HWScan USB hardware inventory"
    workbook.properties.subject = f"Hardware report {report.report_id}"
    workbook.properties.creator = "HWScan USB"
    workbook.save(path)


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    expected = {
        "Summary",
        "Memory",
        "Storage",
        "Graphics",
        "Network",
        "Battery",
        "Displays",
        "Devices",
        "Diagnostics",
        "Metadata",
    }
    missing = expected.difference(workbook.sheetnames)
    workbook.close()
    if missing:
        raise ValueError(f"workbook is missing sheets: {', '.join(sorted(missing))}")


def _append_mapping(sheet: Any, title: str, values: Mapping[str, Any]) -> None:
    sheet.append([title])
    sheet.cell(sheet.max_row, 1).font = Font(bold=True, color="17324D")
    for key, value in _flatten(values):
        sheet.append([key, excel_safe_text(value)])
        sheet.cell(sheet.max_row, 2).number_format = "@" if isinstance(value, str) else "General"
    sheet.append([])


def _flatten(values: Mapping[str, Any], prefix: str = "") -> Iterable[tuple[str, Any]]:
    for key, value in values.items():
        label = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            yield from _flatten(value, label)
        elif isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
            yield label, ", ".join(str(item) for item in value)
        else:
            yield label, value


def _add_rows_sheet(
    workbook: Workbook, name: str, rows: list[dict[str, Any]], table_name: str
) -> None:
    sheet = workbook.create_sheet(name)
    headers = _headers(rows) or ["status"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append([excel_safe_text(row.get(header)) for header in headers])
        for column, value in enumerate((row.get(header) for header in headers), start=1):
            if isinstance(value, str):
                sheet.cell(sheet.max_row, column).number_format = "@"
    if not rows:
        sheet.append(["No data collected"] + [None] * (len(headers) - 1))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if rows:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for index, header in enumerate(headers, start=1):
        longest = max([len(str(header)), *(len(str(row.get(header, ""))) for row in rows)])
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), 60)


def _headers(rows: list[dict[str, Any]]) -> list[str]:
    ordered: list[str] = []
    for row in rows:
        for key in row:
            if key not in ordered:
                ordered.append(key)
    return ordered
