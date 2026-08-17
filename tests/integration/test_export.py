import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from hwscan.application.export_service import ExportService
from hwscan.infrastructure.collectors.fixture import FixtureCollector

FIXTURE = Path("tests/fixtures/dell/latitude-7420")


def test_exports_matching_json_excel_and_manifest(tmp_path: Path) -> None:
    report = FixtureCollector(FIXTURE).collect()
    report.operator_input.notes = "=not-a-formula"
    result = ExportService().export(report, tmp_path)

    data = json.loads(result.json_path.read_text(encoding="utf-8"))
    assert data["report_id"] == report.report_id
    workbook = load_workbook(result.excel_path, read_only=True, data_only=False)
    assert workbook["Summary"]["B2"].value == report.report_id
    assert workbook["Summary"]["B7"].value == "'=not-a-formula"
    assert set(workbook.sheetnames) == {
        "Summary",
        "Memory",
        "Storage",
        "Graphics",
        "Network",
        "Battery",
        "Displays",
        "Devices",
        "Diagnostics",
        "Metadata",
    }
    workbook.close()

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["report_id"] == report.report_id
    for record in manifest["files"]:
        path = tmp_path / record["name"]
        assert record["sha256"] == hashlib.sha256(path.read_bytes()).hexdigest()
    assert not list(tmp_path.glob("*.part"))
